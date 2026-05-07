import re
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

INPUT_ROOT = Path("./data")
TARGET_YEARS = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]

if TARGET_YEARS:
    _min_year = min(TARGET_YEARS)
    _max_year = max(TARGET_YEARS)
    OUTPUT_CSV = Path(f"./clean_data/combined_ge_turbines_{_min_year % 100:02d}_{_max_year % 100:02d}.csv")
else:
    OUTPUT_CSV = Path("./clean_data/combined_ge_turbines.csv")

POSSIBLE_SHEET_TITLES = {
    "ge turbines status",
    "ge turbines",
    "ge turbines sheet",
    "ge turbines status sheet",
}


def normalize(text):
    if pd.isna(text):
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(text).lower()).strip()


def parse_date_from_filename(name):
    match = re.search(r"(\d{4}[-_ ]\d{1,2}[-_ ]\d{1,2}|\d{1,2}[-_ ]\d{1,2}[-_ ]\d{2,4})", name)
    if not match:
        return None

    raw = match.group(1).replace("_", "-").replace(" ", "-")
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m-%d-%Y", "%d-%m-%y", "%m-%d-%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def detect_sheet_name(workbook):
    for sheet_name in workbook.sheetnames:
        if normalize(sheet_name) in POSSIBLE_SHEET_TITLES:
            return sheet_name

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=25, max_col=8, values_only=True):
            row_text = " ".join(normalize(cell) for cell in row if not pd.isna(cell))
            if "ge turbines status" in row_text or "ge turbines" in row_text:
                return sheet_name

    return None


def find_header_row(ws):
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=39, max_col=8, values_only=True), start=1):
        normalized_cells = [normalize(cell) for cell in row]
        joined = " | ".join(normalized_cells)
        next_row = []
        if row_idx < 39:
            next_row = [normalize(cell.value) for cell in ws[row_idx + 1][:8]]

        has_main_labels = (
            "conc" in joined
            and ("units" in joined or "unit" in joined)
            and ("load" in joined or "load mw" in joined)
            and "status" in joined
            and "remarks" in joined
        )
        next_joined = " | ".join(next_row)
        has_sub_labels = "mw" in next_joined and "on" in next_joined and "off" in next_joined

        if has_main_labels and has_sub_labels:
            return row_idx
    return None


def canonical_conc(value):
    text = normalize(value)
    if not text:
        return None
    if "gialo" in text or text == "59e":
        return "GIALO"
    if "waha" in text:
        return "WAHA"
    if "defa" in text:
        return "DEFA"
    if "dahra" in text or text == "32":
        return "DAHRA"
    return None


def find_conc_in_row(cells):
    for cell in cells[:3]:
        conc = canonical_conc(cell)
        if conc:
            return conc
    return None


def is_total_row(row_text):
    text = normalize(row_text)
    return "total" in text or "sub total" in text or "subtotal" in text


def format_date_value(value):
    if pd.isna(value):
        return np.nan
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text if text else np.nan


def clean_unit_name(value, conc_name):
    text = normalize(value)
    if not text:
        return np.nan

    match = re.search(r"ge\s*(?:#|number)?\s*[- ]*\s*(\d+)", text)
    if match and conc_name:
        return f"{conc_name.lower()}-ge-{match.group(1)}"

    if conc_name:
        return f"{conc_name.lower()}-{text.replace(' ', '-')}"

    return text.replace(" ", "-")


def is_ge_unit(value):
    text = normalize(value)
    if not text:
        return False
    return re.search(r"ge\s*(?:#|number)?\s*[- ]*\s*\d+", text) is not None


def process_files():
    raw_files = INPUT_ROOT.rglob("*.xlsx")
    if TARGET_YEARS:
        files = [
            path
            for path in raw_files
            if not path.name.startswith("~$") and any(str(year) in path.name for year in TARGET_YEARS)
        ]
    else:
        files = [path for path in raw_files if not path.name.startswith("~$")]

    files = sorted(files)
    print(f"Found {len(files)} files matching the target years. Starting GE-turbines cleaning...")

    all_rows = []

    for path in files:
        report_date = parse_date_from_filename(path.name)
        if not report_date:
            print(f"Warning: Could not parse date from {path.name}. Skipping.")
            continue

        if TARGET_YEARS and report_date.year not in TARGET_YEARS:
            continue

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            print(f"Error reading {path.name}: {exc}")
            continue

        target_sheet = detect_sheet_name(workbook)
        if not target_sheet:
            print(f"Warning: GE-turbines sheet not found in {path.name}. Sheets: {workbook.sheetnames}")
            continue

        ws = workbook[target_sheet]
        header_row = find_header_row(ws)
        if not header_row:
            print(f"Warning: Could not find the GE-turbines header row in {path.name}.")
            continue

        current_conc = None
        pending_rows = []

        def flush_pending(conc_name):
            if not pending_rows or not conc_name:
                pending_rows.clear()
                return
            for pending in pending_rows:
                unit_name = clean_unit_name(pending["unit"], conc_name)
                if pd.isna(unit_name):
                    continue
                all_rows.append(
                    {
                        "date": report_date,
                        "conc": conc_name,
                        "unit": unit_name,
                        "load_mw": pd.to_numeric(pending["load"], errors="coerce"),
                        "status": pending["status"],
                        "date_on": format_date_value(pending["date_on"]),
                        "date_off": format_date_value(pending["date_off"]),
                        "remarks": pending["remarks"],
                    }
                )
            pending_rows.clear()

        for row in ws.iter_rows(min_row=header_row + 2, max_col=8, values_only=True):
            cells = list(row)
            row_text = " ".join(normalize(cell) for cell in cells if not pd.isna(cell))

            if not row_text:
                continue
            if "general remarks" in row_text:
                break
            if is_total_row(row_text):
                current_conc = None
                pending_rows.clear()
                continue

            unit_value = cells[1] if len(cells) > 1 else np.nan
            load_value = cells[2] if len(cells) > 2 else np.nan
            status_value = cells[4] if len(cells) > 4 else np.nan
            date_on_value = cells[5] if len(cells) > 5 else np.nan
            date_off_value = cells[6] if len(cells) > 6 else np.nan
            remarks_value = cells[7] if len(cells) > 7 else np.nan
            status_text = str(status_value).strip() if not pd.isna(status_value) else np.nan
            remarks_text = str(remarks_value).strip() if not pd.isna(remarks_value) else np.nan

            potential_conc = find_conc_in_row(cells)
            if potential_conc:
                current_conc = potential_conc
                flush_pending(current_conc)
                unit_is_conc = canonical_conc(unit_value) is not None
                if pd.isna(unit_value) or not str(unit_value).strip() or unit_is_conc:
                    continue

            if canonical_conc(unit_value) is not None:
                current_conc = canonical_conc(unit_value)
                flush_pending(current_conc)
                continue

            if pd.isna(unit_value) or not str(unit_value).strip():
                continue

            if current_conc is None:
                if is_ge_unit(unit_value):
                    pending_rows.append(
                        {
                            "unit": unit_value,
                            "load": load_value,
                            "status": status_text,
                            "date_on": date_on_value,
                            "date_off": date_off_value,
                            "remarks": remarks_text,
                        }
                    )
                continue

            unit_name = clean_unit_name(unit_value, current_conc)
            if pd.isna(unit_name):
                continue

            all_rows.append(
                {
                    "date": report_date,
                    "conc": current_conc,
                    "unit": unit_name,
                    "load_mw": pd.to_numeric(load_value, errors="coerce"),
                    "status": status_text,
                    "date_on": format_date_value(date_on_value),
                    "date_off": format_date_value(date_off_value),
                    "remarks": remarks_text,
                }
            )

    if not all_rows:
        print("No GE-turbines data processed.")
        return

    combined = pd.DataFrame(all_rows)
    combined = combined.dropna(subset=["conc", "unit"]).copy()
    combined = combined.sort_values(["date", "conc", "unit"]).reset_index(drop=True)
    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    combined.to_csv(OUTPUT_CSV, index=False)

    print(f"Success! Wrote {len(combined)} rows to {OUTPUT_CSV}")


if __name__ == "__main__":
    process_files()